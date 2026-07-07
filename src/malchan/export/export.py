import numpy as np
import pandas as pd
from matplotlib import cm
from matplotlib.colors import Normalize
from statsmodels.stats.outliers_influence import variance_inflation_factor
from sklearn.feature_selection import mutual_info_regression

import matplotlib.pyplot as plt

import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import io
import os
import sys
sys.path.append('..')

from machine_learning.models.export.export_visualization import (
    save_yy_plot_mpl, save_importances_plot_mpl, plot_shap_beeswarm_mpl, save_pd_plot_mpl, plot_pd_heatmap_matrix_lower,
    save_pca_mpl, save_corrmatrix_mpl, save_scattermatrix_mpl
)

import warnings
warnings.simplefilter('ignore')

IMAGE_PATH = os.path.join(os.path.dirname(__file__), "images")
# IMAGE_PATH = "machine_learning/models/visualization/images"

def get_excel_file(workbook):
    buffer = io.BytesIO()
    workbook.save(buffer)  # ワークブックをバイト形式で保存
    buffer.seek(0)  # バッファの先頭に戻す

    # 古い画像を削除
    img_files = os.listdir(IMAGE_PATH)
    for img in img_files:
        os.remove(IMAGE_PATH+'/' + img)
    return buffer

def save_mpl_images(model):
    for target in model.target_cols:
        save_yy_plot_mpl(model, target)
        save_importances_plot_mpl(model, target)
        plot_shap_beeswarm_mpl(model, target)
        save_pd_plot_mpl(model, target)
        if model.models[target].target_items is not None:
            for class_label in model.models[target].target_items:
                plot_pd_heatmap_matrix_lower(model, target, class_label=class_label)
        else:
            plot_pd_heatmap_matrix_lower(model, target)
        save_pca_mpl(model, target)
    
    save_corrmatrix_mpl(model)
    save_scattermatrix_mpl(model)

def write_cv_scores(model, target, ws):
    score_types = model.models[target].cv_scores["train"].columns
    train_values = model.models[target].cv_scores["train"].values.ravel()
    test_values = model.models[target].cv_scores["test"].values.ravel()

    ws["Q1"] = "精度(交差検証)"
    ws["R2"] = score_types[0]
    ws["S2"] = score_types[1]
    ws["T2"] = score_types[2]
    ws["U2"] = score_types[3]
    
    ws["Q3"] = "train"
    ws["R3"] = train_values[0]
    ws["S3"] = train_values[1]
    ws["T3"] = train_values[2]
    ws["U3"] = train_values[3]

    ws["Q4"] = "test"
    ws["R4"] = test_values[0]
    ws["S4"] = test_values[1]
    ws["T4"] = test_values[2]
    ws["U4"] = test_values[3]

def write_parameters(model, target, wb):
    # target = model.models[target].target_cols[0]
    ws = wb.create_sheet(title="モデル設定("+target+")")
    
    alphabet_list = [chr(i) for i in range(65, 91)]
    
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    
    ws['A1'] = "モデル名"
    for i in range(len(model.models[target].model_names)):
        ws[alphabet_list[i+1]+"1"] = model.models[target].model_names[i]

    ws['A2'] = "タスク"
    ws['B2'] = model.models[target].task
    
    ws['A3'] = "チューニング"
    ws['B3'] = "True" if model.models[target].tuning else "False"
    
    ws['A4'] = "アンサンブル"
    ws['B4'] = "True" if model.models[target].ensemble else "False"
    
    ws['A5'] = "アンサンブル手法"
    ws['B5'] = model.models[target].ens_type if model.models[target].ensemble else ""
    
    ws['A6'] = "ベースモデル"
    ws['B6'] =model.models[target].base_model if (model.models[target].ensemble)&(model.models[target].ens_type!="アンサンブル") else ""
    
    ws['A7'] = "補完(連続値)"
    ws['B7'] = model.models[target].num_impute_type
    
    ws['A8'] = "正規化(連続値)"
    ws['B8'] = model.models[target].num_scale_type
    
    ws['A9'] = "補完(カテゴリ値)"
    ws['B9'] = "True" if model.models[target].cat_impute else "False"
    
    ws['A10'] = "多項式特徴量"
    ws['B10'] = "True" if model.models[target].poly else "False"
    
    ws['A11'] = "多項式次元"
    ws['B11'] = model.models[target].poly if model.models[target].poly else ""
    
    ws['A11'] = "交互作用項のみ"
    if model.models[target].poly:
        if model.models[target].poly_interaction_only:
            ws['B12'] = "True"
        else:
            ws['B12'] = "False"
    else:
        ws['B12'] = ""
    
    ws['A13'] = "次元削減"
    ws['B13'] = "True" if model.models[target].decomposition else "False"
    
    ws['A14'] = "手法/削減数"
    ws['B14'] = model.models[target].decomposition_method if model.models[target].decomposition else ""
    ws['C14'] = model.models[target].dec_n_components if model.models[target].decomposition else ""
    
    # すべてのパラメータを取得
    params = model.models[target].model.get_params()
    
    # 数値のパラメータのみを取得
    num_params = {k: [v] for k, v in params.items() if isinstance(v, (int, float))}
    
    num_params = pd.DataFrame(num_params)
    num_params = num_params.filter(like="predictor")
    num_params = num_params.rename(columns=lambda x: x.replace("predictor__", ""))
    num_params = num_params.T
    
    ws['A16'] = "パラメータ"
    for i in range(len(num_params)):
        ws['A'+str(int(17+i))] = num_params.index[i]
        ws['B'+str(int(17+i))] = num_params[0][i]
        last_idx = 18+i

def make_sheet(model):
    save_mpl_images(model)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '散布図行列'
    
    # 3. 画像をExcelシートに挿入
    graph_image = f"{IMAGE_PATH}/scattermatrix.png"
    img = Image(graph_image)
    ws.add_image(img, 'A1')  # 'A1'のセルに画像を挿入
    
    ws = wb.create_sheet(title="相関行列")
    graph_image = f"{IMAGE_PATH}/corrmatrix.png"
    img = Image(graph_image)
    ws.add_image(img, 'A1')  # 'A1'のセルに画像を挿入

    for target in model.target_cols:
        write_parameters(model, target, wb)
        ws = wb.create_sheet(title="機械学習結果("+target+")")
        graph_image = f"{IMAGE_PATH}/yyplot_image_"+target+".png"
        img = Image(graph_image)
        ws.add_image(img, 'C1')  # 'A1'のセルに画像を挿入

        write_cv_scores(model, target, ws)
        
        graph_image = f"{IMAGE_PATH}/importances_"+target+".png"
        img = Image(graph_image)
        ws.add_image(img, 'A25')  # 'A1'のセルに画像を挿入
        
        graph_image = f"{IMAGE_PATH}/beeswarm_"+target+".png"
        img = Image(graph_image)
        ws.add_image(img, 'N25')  # 'A1'のセルに画像を挿入
        
        graph_image = f"{IMAGE_PATH}/pd_"+target+".png"
        img = Image(graph_image)
        ws.add_image(img, 'A60')  # 'A1'のセルに画像を挿入

        if model.models[target].target_items is not None:
            next_idx = 60 + 18*len(model.models[target].target_items)
            for class_label in model.models[target].target_items: 
                graph_image = f"{IMAGE_PATH}/pd2d_{target}_{str(class_label)}.png"
                img = Image(graph_image)
                ws.add_image(img, f'A{str(next_idx)}')  # 'A1'のセルに画像を挿入
                next_idx += 15 * len(model.models[target].all_cols)
        else:
            next_idx = 60 + 18

            graph_image = f"{IMAGE_PATH}/pd2d_"+target+".png"
            img = Image(graph_image)
            ws.add_image(img, f'A{str(next_idx)}')  # 'A1'のセルに画像を挿入
        
        if os.path.exists(f"{IMAGE_PATH}/dec_"+target+".png"):
            graph_image = f"{IMAGE_PATH}/dec_"+target+".png"
            img = Image(graph_image)
            ws.add_image(img, 'W1')  # 'A1'のセルに画像を挿入

    return get_excel_file(wb)
    # return wb